"""多 Sheet Excel 导出器：精确复刻已验证样本格式
- 一般纳税人：4 Sheet（发票汇总、航空行程单明细、抵扣汇总（按税率归类）、发票汇总-费用归类）
- 小规模纳税人：2 Sheet（发票明细、发票汇总-费用归类）
"""
from __future__ import annotations
from typing import TYPE_CHECKING, Any
from collections import OrderedDict

if TYPE_CHECKING:
    from core.models import InvoiceRecord, TaxpayerType


class ExcelExporter:
    """精确复刻已验证样本格式的多 Sheet 导出器"""

    _openpyxl = None
    _styles = None

    @classmethod
    def _ensure_styles(cls):
        if cls._styles is not None:
            return cls._styles
        import openpyxl
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        cls._openpyxl = openpyxl
        cls._openpyxl_utils = openpyxl.utils
        cls._Alignment = Alignment
        cls._styles = {
            "HEADER_FONT": Font(bold=True, size=11),
            "HEADER_FILL": PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid"),
            "HEADER_ALIGN": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "BOLD_FONT": Font(bold=True, size=11),
            "NORMAL_FONT": Font(size=11),
            "THIN_BORDER": Border(
                left=Side(style="thin", color="D9D9D9"),
                right=Side(style="thin", color="D9D9D9"),
                top=Side(style="thin", color="D9D9D9"),
                bottom=Side(style="thin", color="D9D9D9"),
            ),
            "SUMMARY_FILL": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
            "SUMMARY_FONT": Font(bold=True, size=11),
            "NUMBER_FMT": "#,##0.00",
        }
        return cls._styles

    def __init__(self, taxpayer_type):
        self.taxpayer_type = taxpayer_type

    def _write_header(self, ws, headers):
        r = self._styles
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.font = r["HEADER_FONT"]
            cell.fill = r["HEADER_FILL"]
            cell.alignment = r["HEADER_ALIGN"]
            cell.border = r["THIN_BORDER"]

    def _write_row(self, ws, row_idx, row_data, fmt_align, bold=False, fill=None):
        r = self._styles
        font = r["BOLD_FONT"] if bold else r["NORMAL_FONT"]
        for col_idx, (val, (num_fmt, align)) in enumerate(zip(row_data, fmt_align), 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = font
            cell.number_format = num_fmt
            cell.alignment = self._Alignment(horizontal=align, vertical="center")
            cell.border = r["THIN_BORDER"]
            if fill:
                cell.fill = fill

    def _set_widths(self, ws, widths):
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[self._openpyxl_utils.get_column_letter(i)].width = w

    def _fmt_align(self, *specs):
        """Build format+alignment list: (num_fmt, align)"""
        return list(specs)

    # ────────── 一般纳税人 4 Sheet ──────────

    def _build_general_sheets(self, wb, records):
        r = self._styles

        # ===== 统一排序规则 =====
        sorted_records = sorted(records, key=lambda rec: (
            (rec.item_name or "").strip(),
            (rec.seller_name or "").strip(),
        ))

        # ─────────────────────────────────────────
        # Sheet 1: 发票汇总
        # ─────────────────────────────────────────
        ws = wb.active
        ws.title = "发票汇总"
        headers = [
            "序号", "发票类型", "发票号码", "开票日期", "销售方名称", "项目名称",
            "金额（不含税）", "税率", "税额", "价税合计", "可抵扣税额", "源文件",
        ]
        widths = [6, 18, 22, 16, 38, 32, 16, 10, 14, 16, 16, 50]
        fmt = self._fmt_align(
            ("General", "center"),    # 序号
            ("General", "left"),      # 发票类型
            ("General", "center"),    # 发票号码
            ("General", "center"),    # 开票日期
            ("General", "left"),      # 销售方名称
            ("General", "left"),      # 项目名称
            ("#,##0.00", "right"),    # 金额（不含税）
            ("General", "center"),    # 税率
            ("#,##0.00", "right"),    # 税额
            ("#,##0.00", "right"),    # 价税合计
            ("#,##0.00", "right"),    # 可抵扣税额
            ("General", "left"),      # 源文件
        )
        self._write_header(ws, headers)
        self._set_widths(ws, widths)

        seq = 0
        for rec in sorted_records:
            seq += 1
            row_data = [
                seq, rec.invoice_type, rec.invoice_number, rec.invoice_date,
                rec.seller_name, rec.item_name,
                round(rec.amount_no_tax, 2), rec.tax_rate,
                round(rec.tax_amount, 2), round(rec.total_amount, 2),
                round(rec.deductible_tax, 2), rec.source_file,
            ]
            self._write_row(ws, seq + 1, row_data, fmt, bold=rec.deductible_tax > 0.005)

        # 合计行
        total_row = seq + 2
        totals = {
            "nontax": sum(round(r.amount_no_tax, 2) for r in sorted_records),
            "tax": sum(round(r.tax_amount, 2) for r in sorted_records),
            "total": sum(round(r.total_amount, 2) for r in sorted_records),
            "deduct": sum(round(r.deductible_tax, 2) for r in sorted_records),
        }
        total_data = [
            "合计", "", "", "", "", "",
            round(totals["nontax"], 2), "",
            round(totals["tax"], 2), round(totals["total"], 2),
            round(totals["deduct"], 2), "",
        ]
        self._write_row(ws, total_row, total_data, fmt, bold=True, fill=r["SUMMARY_FILL"])
        # 隐藏源文件列（第12列=L）
        ws.column_dimensions["L"].hidden = True

        # ─────────────────────────────────────────
        # Sheet 2: 航空行程单明细
        # ─────────────────────────────────────────
        ws2 = wb.create_sheet("航空行程单明细")
        air_records = [rec for rec in sorted_records
                       if rec.invoice_type and "航空运输电子客票行程单" in rec.invoice_type]
        if not air_records:
            ws2.cell(row=1, column=1, value="无航空行程单数据")
            ws2.cell(row=1, column=1).font = r["NORMAL_FONT"]
        else:
            air_headers = [
                "发票号码", "旅客姓名", "证件号码", "航程简述", "航段数",
                "票价", "民航发展基金", "燃油附加费", "其他税费",
                "合计", "可抵扣税额", "填开单位", "填开日期", "源文件",
            ]
            air_widths = [22, 16, 22, 28, 8, 14, 14, 14, 12, 14, 14, 28, 16, 50]
            air_fmt = self._fmt_align(
                ("General", "center"),    # 发票号码
                ("General", "center"),    # 旅客姓名
                ("General", "center"),    # 证件号码
                ("General", "left"),      # 航程简述
                ("General", "center"),    # 航段数
                ("#,##0.00", "right"),    # 票价
                ("#,##0.00", "right"),    # 民航发展基金
                ("#,##0.00", "right"),    # 燃油附加费
                ("#,##0.00", "right"),    # 其他税费
                ("#,##0.00", "right"),    # 合计
                ("#,##0.00", "right"),    # 可抵扣税额
                ("General", "left"),      # 填开单位
                ("General", "center"),    # 填开日期
                ("General", "left"),      # 源文件
            )
            self._write_header(ws2, air_headers)
            self._set_widths(ws2, air_widths)
            air_totals = {"票价": 0.0, "CAF": 0.0, "燃油": 0.0, "其他": 0.0, "合计": 0.0, "可抵扣": 0.0}
            for i, rec in enumerate(air_records):
                ext = rec.raw_data or {}
                ticket_price = float(ext.get("票价", 0) or 0)
                civil_fund = float(ext.get("民航发展基金", 0) or 0)
                fuel = float(ext.get("燃油附加费", 0) or 0)
                other = float(ext.get("其他税费", 0) or 0)
                tot = round(rec.total_amount, 2)
                deduct = round(rec.deductible_tax, 2)

                row_data = [
                    rec.invoice_number,
                    ext.get("旅客姓名", ""),
                    ext.get("证件号码", ""),
                    ext.get("航程简述", ""),
                    ext.get("航段数", 1),
                    round(ticket_price, 2),
                    round(civil_fund, 2),
                    round(fuel, 2),
                    round(other, 2),
                    tot, deduct,
                    ext.get("填开单位", rec.seller_name),
                    ext.get("填开日期", rec.invoice_date),
                    rec.source_file,
                ]
                self._write_row(ws2, i + 2, row_data, air_fmt, bold=rec.deductible_tax > 0.005)
                air_totals["票价"] += ticket_price
                air_totals["CAF"] += civil_fund
                air_totals["燃油"] += fuel
                air_totals["其他"] += other
                air_totals["合计"] += tot
                air_totals["可抵扣"] += deduct

            # 合计行
            a_total_row = len(air_records) + 2
            a_total_data = [
                "合计", "", "", "", "",
                round(air_totals["票价"], 2), round(air_totals["CAF"], 2),
                round(air_totals["燃油"], 2), round(air_totals["其他"], 2),
                round(air_totals["合计"], 2), round(air_totals["可抵扣"], 2),
                "", "", "",
            ]
            self._write_row(ws2, a_total_row, a_total_data, air_fmt, bold=True, fill=r["SUMMARY_FILL"])
            ws2.column_dimensions["N"].hidden = True

        # ─────────────────────────────────────────
        # Sheet 3: 抵扣汇总（按税率归类）
        # ─────────────────────────────────────────
        ws3 = wb.create_sheet("抵扣汇总（按税率归类）")
        deduct_headers = ["发票类型", "份数", "金额合计", "税额合计"]
        deduct_widths = [28, 10, 16, 16]
        deduct_fmt = self._fmt_align(
            ("General", "left"),      # 发票类型
            ("General", "center"),    # 份数
            ("#,##0.00", "right"),    # 金额合计
            ("#,##0.00", "right"),    # 税额合计
        )
        self._write_header(ws3, deduct_headers)
        self._set_widths(ws3, deduct_widths)

        # 按发票类型分组（仅可抵扣的记录）
        deduct_records = [rec for rec in sorted_records if rec.deductible_tax > 0.005]
        type_groups = OrderedDict()
        for rec in deduct_records:
            key = str(rec.invoice_type or "其他")
            group = type_groups.get(key, {"count": 0, "nontax": 0.0, "tax": 0.0})
            group["count"] += 1
            group["nontax"] += round(rec.amount_no_tax, 2)
            group["tax"] += round(rec.deductible_tax, 2)
            type_groups[key] = group

        d_row = 2
        grand_count = 0
        grand_nontax = 0.0
        grand_tax = 0.0
        for inv_type, group in type_groups.items():
            row_data = [
                inv_type, group["count"],
                round(group["nontax"], 2), round(group["tax"], 2),
            ]
            self._write_row(ws3, d_row, row_data, deduct_fmt)
            grand_count += group["count"]
            grand_nontax += group["nontax"]
            grand_tax += group["tax"]
            d_row += 1

        # 合计行
        d_total_data = ["合计", grand_count, round(grand_nontax, 2), round(grand_tax, 2)]
        self._write_row(ws3, d_row, d_total_data, deduct_fmt, bold=True, fill=r["SUMMARY_FILL"])

        # ─────────────────────────────────────────
        # Sheet 4: 发票汇总-费用归类
        # ─────────────────────────────────────────
        from core.category_mapper import classify_item

        ws4 = wb.create_sheet("发票汇总-费用归类")
        cat_headers = [
            "管理费用二级明细科目", "发票号码", "开票日期", "销售方名称", "项目名称",
            "金额（不含税）", "税额", "价税合计", "可抵扣税额",
        ]
        cat_widths = [22, 22, 16, 38, 32, 16, 14, 16, 16]
        cat_fmt = self._fmt_align(
            ("General", "left"),      # 管理费用二级明细科目
            ("General", "center"),    # 发票号码
            ("General", "center"),    # 开票日期
            ("General", "left"),      # 销售方名称
            ("General", "left"),      # 项目名称
            ("#,##0.00", "right"),    # 金额（不含税）
            ("#,##0.00", "right"),    # 税额
            ("#,##0.00", "right"),    # 价税合计
            ("#,##0.00", "right"),    # 可抵扣税额
        )
        self._write_header(ws4, cat_headers)
        self._set_widths(ws4, cat_widths)

        # 费用归类顺序（与已验证样本一致）
        CAT_ORDER = ["差旅费", "市内交通费", "业务招待费", "办公费", "燃料费", "维修费", "车辆使用费", "其他费用"]

        def get_category(rec):
            cat = classify_item(rec.item_name or "", rec.seller_name or "")
            if not cat or cat == "其他":
                return "其他费用"
            return cat

        # 按类别分组
        cat_groups = OrderedDict((c, []) for c in CAT_ORDER)
        for rec in sorted_records:
            cat = get_category(rec)
            if cat not in cat_groups:
                cat_groups[cat] = []
            cat_groups[cat].append(rec)

        c_row = 2
        grand_total = {"count": 0, "nontax": 0.0, "tax": 0.0, "total": 0.0, "deduct": 0.0}
        for cat, cat_records in cat_groups.items():
            if not cat_records:
                continue
            # 类别内按(item_name, seller_name)排序
            cat_records.sort(key=lambda rec: (
                (rec.item_name or "").strip(),
                (rec.seller_name or "").strip(),
            ))
            cat_sub = {"count": 0, "nontax": 0.0, "tax": 0.0, "total": 0.0, "deduct": 0.0}
            for rec in cat_records:
                row_data = [
                    cat, rec.invoice_number, rec.invoice_date,
                    rec.seller_name, rec.item_name,
                    round(rec.amount_no_tax, 2),
                    round(rec.tax_amount, 2),
                    round(rec.total_amount, 2),
                    round(rec.deductible_tax, 2),
                ]
                self._write_row(ws4, c_row, row_data, cat_fmt, bold=rec.deductible_tax > 0.005)
                cat_sub["count"] += 1
                cat_sub["nontax"] += round(rec.amount_no_tax, 2)
                cat_sub["tax"] += round(rec.tax_amount, 2)
                cat_sub["total"] += round(rec.total_amount, 2)
                cat_sub["deduct"] += round(rec.deductible_tax, 2)
                c_row += 1

            # 小计行
            sub_row = ["  【{}】小计 共{}笔".format(cat, cat_sub["count"]),
                       "", "", "", "",
                       round(cat_sub["nontax"], 2),
                       round(cat_sub["tax"], 2),
                       round(cat_sub["total"], 2),
                       round(cat_sub["deduct"], 2)]
            self._write_row(ws4, c_row, sub_row, cat_fmt, bold=True, fill=r["SUMMARY_FILL"])
            c_row += 1

            grand_total["count"] += cat_sub["count"]
            grand_total["nontax"] += cat_sub["nontax"]
            grand_total["tax"] += cat_sub["tax"]
            grand_total["total"] += cat_sub["total"]
            grand_total["deduct"] += cat_sub["deduct"]

        # 总计行
        grand_row = [
            "  【总计】共{}笔".format(grand_total["count"]),
            "", "", "", "",
            round(grand_total["nontax"], 2),
            round(grand_total["tax"], 2),
            round(grand_total["total"], 2),
            round(grand_total["deduct"], 2),
        ]
        self._write_row(ws4, c_row, grand_row, cat_fmt, bold=True, fill=r["SUMMARY_FILL"])

    # ────────── 小规模纳税人 2 Sheet ──────────

    SMALL_HEADERS = ["序号", "发票类型", "发票号码", "开票日期", "销售方名称", "项目名称", "价税合计（小写）"]
    SMALL_WIDTHS = [6, 18, 22, 16, 38, 32, 16]
    SMALL_FMT_ALIGN = [
        ("General", "center"), ("General", "center"),
        ("General", "center"), ("General", "center"),
        ("General", "left"), ("General", "left"),
        ("#,##0.00", "right"),
    ]

    def _build_small_sheets(self, wb, records):
        r = self._styles
        # ---- Sheet 1: 发票明细 ----
        ws = wb.active
        ws.title = "发票明细"
        self._write_header(ws, self.SMALL_HEADERS)
        self._set_widths(ws, self.SMALL_WIDTHS)
        for i, rec in enumerate(records):
            row_data = [
                i + 1, rec.invoice_type, rec.invoice_number,
                rec.invoice_date, rec.seller_name, rec.item_name,
                round(rec.total_amount, 2),
            ]
            self._write_row(ws, i + 2, row_data, self.SMALL_FMT_ALIGN)
        # 合计行
        total_row = len(records) + 2
        total_amt = sum(round(r.total_amount, 2) for r in records)
        total_data = ["合计", "", "", "", "", "", round(total_amt, 2)]
        self._write_row(ws, total_row, total_data, self.SMALL_FMT_ALIGN, bold=True, fill=r["SUMMARY_FILL"])

        # ---- Sheet 2: 发票汇总-费用归类（简版） ----
        from core.category_mapper import classify_item
        ws2 = wb.create_sheet("发票汇总-费用归类")
        cat_headers = [
            "管理费用二级明细科目", "发票号码", "开票日期", "销售方名称", "项目名称",
            "金额（不含税）", "税额", "价税合计", "可抵扣税额",
        ]
        cat_widths = [22, 22, 16, 38, 32, 16, 14, 16, 16]
        cat_fmt = [
            ("General", "left"), ("General", "center"), ("General", "center"),
            ("General", "left"), ("General", "left"),
            ("#,##0.00", "right"), ("#,##0.00", "right"),
            ("#,##0.00", "right"), ("#,##0.00", "right"),
        ]
        self._write_header(ws2, cat_headers)
        self._set_widths(ws2, cat_widths)

        CAT_ORDER = ["差旅费", "市内交通费", "业务招待费", "办公费", "燃料费", "维修费", "车辆使用费", "其他费用"]
        cat_groups = OrderedDict((c, []) for c in CAT_ORDER)
        for rec in records:
            cat = classify_item(rec.item_name or "", rec.seller_name or "")
            if not cat or cat == "其他":
                cat = "其他费用"
            if cat not in cat_groups:
                cat_groups[cat] = []
            cat_groups[cat].append(rec)

        c_row = 2
        grand_total = {"count": 0, "total": 0.0}
        for cat, cat_records in cat_groups.items():
            if not cat_records:
                continue
            cat_sub = {"count": 0, "total": 0.0}
            for rec in cat_records:
                row_data = [
                    cat, rec.invoice_number, rec.invoice_date,
                    rec.seller_name, rec.item_name, "", "",
                    round(rec.total_amount, 2), "",
                ]
                self._write_row(ws2, c_row, row_data, cat_fmt)
                cat_sub["count"] += 1
                cat_sub["total"] += round(rec.total_amount, 2)
                c_row += 1

            sub_row = ["  【{}】小计 共{}笔".format(cat, cat_sub["count"]),
                       "", "", "", "", "", "",
                       round(cat_sub["total"], 2), ""]
            self._write_row(ws2, c_row, sub_row, cat_fmt, bold=True, fill=r["SUMMARY_FILL"])
            c_row += 1
            grand_total["count"] += cat_sub["count"]
            grand_total["total"] += cat_sub["total"]

        grand_row = ["  【总计】共{}笔".format(grand_total["count"]),
                     "", "", "", "", "", "",
                     round(grand_total["total"], 2), ""]
        self._write_row(ws2, c_row, grand_row, cat_fmt, bold=True, fill=r["SUMMARY_FILL"])

    # ────────── 主入口 ──────────

    def export(self, records: list, output_path: str) -> str:
        """导出 Excel"""
        self._ensure_styles()
        wb = self._openpyxl.Workbook()

        if self.taxpayer_type.value == "general":
            self._build_general_sheets(wb, records)
        else:
            self._build_small_sheets(wb, records)

        wb.save(output_path)
        return str(output_path)
